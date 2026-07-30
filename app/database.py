import os
import threading
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl import Workbook

EXCEL_DB_FILE = Path(__file__).resolve().parent.parent / "users_db.xlsx"
lock = threading.Lock()

def get_excel_workbook(file_path):
    """
    Safely opens or creates the workbook.
    """
    if not os.path.exists(file_path):
        wb = Workbook()
        # Initialize Users sheet
        ws_users = wb.active
        ws_users.title = "Users"
        ws_users.append(["Mobile Number", "PIN", "Name", "Email", "Registration Date", "Last Login", "Status"])
        
        # Initialize OTPs sheet
        ws_otps = wb.create_sheet(title="OTPs")
        ws_otps.append(["Email", "OTP", "Expires At"])
        
        wb.save(file_path)
        return wb
    else:
        return openpyxl.load_workbook(file_path)

def init_db():
    """
    Initializes the Excel spreadsheet database.
    """
    print(f"Initializing Excel Database at: {EXCEL_DB_FILE}")
    with lock:
        wb = get_excel_workbook(EXCEL_DB_FILE)
        wb.save(EXCEL_DB_FILE)
        wb.close()
    print("Excel Database initialized successfully.")

def create_user(mobile, pin_hash, email=None, full_name=None):
    """
    Appends a new user row to the Users sheet.
    """
    with lock:
        wb = get_excel_workbook(EXCEL_DB_FILE)
        ws = wb["Users"]
        
        # Verify uniqueness of mobile number (fail-safe check)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == mobile:
                wb.close()
                raise Exception("Mobile number already registered")
        
        reg_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.append([
            str(mobile),
            str(pin_hash),
            str(full_name or ""),
            str(email or ""),
            reg_date,
            "",  # Last Login (empty initially)
            "Active"  # Status
        ])
        
        wb.save(EXCEL_DB_FILE)
        wb.close()
    return True

def get_user_by_mobile(mobile):
    """
    Finds a user in the Users sheet by mobile number.
    """
    with lock:
        if not os.path.exists(EXCEL_DB_FILE):
            return None
        wb = openpyxl.load_workbook(EXCEL_DB_FILE, data_only=True)
        ws = wb["Users"]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Check mobile matches (as string)
            if row[0] and str(row[0]).strip() == str(mobile).strip():
                user_dict = {
                    "mobile": str(row[0]),
                    "pin_hash": str(row[1]),
                    "full_name": str(row[2] or ""),
                    "email": str(row[3] or ""),
                    "created_at": str(row[4] or ""),
                    "last_login": str(row[5] or ""),
                    "status": str(row[6] or "Active")
                }
                wb.close()
                return user_dict
        wb.close()
    return None

def get_user_by_email(email):
    """
    Finds a user in the Users sheet by email address.
    """
    if not email:
        return None
    with lock:
        if not os.path.exists(EXCEL_DB_FILE):
            return None
        wb = openpyxl.load_workbook(EXCEL_DB_FILE, data_only=True)
        ws = wb["Users"]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[3] and str(row[3]).strip().lower() == str(email).strip().lower():
                user_dict = {
                    "mobile": str(row[0]),
                    "pin_hash": str(row[1]),
                    "full_name": str(row[2] or ""),
                    "email": str(row[3] or ""),
                    "created_at": str(row[4] or ""),
                    "last_login": str(row[5] or ""),
                    "status": str(row[6] or "Active")
                }
                wb.close()
                return user_dict
        wb.close()
    return None

def update_last_login(mobile):
    """
    Updates the Last Login timestamp for the user in the Users sheet.
    """
    with lock:
        if not os.path.exists(EXCEL_DB_FILE):
            return
        wb = openpyxl.load_workbook(EXCEL_DB_FILE)
        ws = wb["Users"]
        
        login_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        updated = False
        
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] and str(row[0]).strip() == str(mobile).strip():
                # Write to column 6 (Last Login)
                ws.cell(row=idx, column=6, value=login_date)
                updated = True
                break
        
        if updated:
            wb.save(EXCEL_DB_FILE)
        wb.close()

def store_otp(email, otp, expires_at):
    """
    Stores or replaces an OTP record in the OTPs sheet.
    """
    with lock:
        wb = get_excel_workbook(EXCEL_DB_FILE)
        ws = wb["OTPs"]
        
        # Check if record already exists, if so overwrite it
        row_found = None
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] and str(row[0]).strip().lower() == str(email).strip().lower():
                row_found = idx
                break
                
        if row_found:
            ws.cell(row=row_found, column=2, value=str(otp))
            ws.cell(row=row_found, column=3, value=int(expires_at))
        else:
            ws.append([
                str(email),
                str(otp),
                int(expires_at)
            ])
            
        wb.save(EXCEL_DB_FILE)
        wb.close()

def get_otp(email):
    """
    Finds an OTP record in the OTPs sheet.
    """
    with lock:
        if not os.path.exists(EXCEL_DB_FILE):
            return None
        wb = openpyxl.load_workbook(EXCEL_DB_FILE, data_only=True)
        ws = wb["OTPs"]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and str(row[0]).strip().lower() == str(email).strip().lower():
                otp_dict = {
                    "email": str(row[0]),
                    "otp": str(row[1]),
                    "expires_at": int(row[2])
                }
                wb.close()
                return otp_dict
        wb.close()
    return None

def delete_otp(email):
    """
    Deletes an OTP record from the OTPs sheet.
    """
    with lock:
        if not os.path.exists(EXCEL_DB_FILE):
            return
        wb = openpyxl.load_workbook(EXCEL_DB_FILE)
        ws = wb["OTPs"]
        
        row_to_delete = None
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] and str(row[0]).strip().lower() == str(email).strip().lower():
                row_to_delete = idx
                break
                
        if row_to_delete:
            ws.delete_rows(row_to_delete, 1)
            wb.save(EXCEL_DB_FILE)
        wb.close()
